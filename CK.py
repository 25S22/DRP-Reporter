"""
=============================================================================
INCIDENT REPORT GENERATOR
=============================================================================
CONFIGURATION (only 3 things to set):
  INPUT_FILE_PATH  — path to the source Excel workbook
  COL_INCIDENT_ID  — exact column name for the Incident ID
  COL_CLOSURE_DATE — exact column name for the Closure Date

Everything else — date range (prompted at runtime), number of sheets,
number of modules with data, chart colours, layout — is handled automatically.

Date formats like "2 Mar 2026", "2nd March 2026", "03/02/2026", "2026-03-02"
are all parsed correctly.
=============================================================================
"""

import os
import re
import sys
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn, CustomFilter, CustomFilters
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# >>>  CONFIGURATION — only edit these three lines  <<<
# =============================================================================

INPUT_FILE_PATH  = "incidents.xlsx"
COL_INCIDENT_ID  = "Incident Id"
COL_CLOSURE_DATE = "Incident Closure on"

# =============================================================================
# INTERNALS — do not edit below this line
# =============================================================================

OUTPUT_FILE_PATH   = "incidents_report.xlsx"
SUMMARY_SHEET_NAME = "Summary Dashboard"

# Colour palette — extended so any number of modules gets a unique colour.
# Cycles if there are more modules than colours (graceful, never errors).
_PALETTE = [
    "4472C4","ED7D31","70AD47","FF6961","FFC000","9DC3E6","A9D18E",
    "FF9DA7","7030A0","00B0F0","F4B942","C00000","375623","833C00",
    "5B9BD5","E2EFDA","BDD7EE","FCE4D6","EDEDED","002060","843C0C",
    "1F497D","60497A","215732","7F7F7F","D9E1F2","F8CBAD","E2EFDA",
]

def _color(idx):
    return _PALETTE[idx % len(_PALETTE)]

# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    m = Side(style="medium")
    return Border(left=m, right=m, top=m, bottom=m)

NAVY       = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F7"
ALT_BLUE   = "EFF5FB"
WHITE      = "FFFFFF"

HEADER_FILL = PatternFill("solid", start_color=NAVY)
SUBHDR_FILL = PatternFill("solid", start_color=MID_BLUE)
TOTAL_FILL  = PatternFill("solid", start_color=LIGHT_BLUE)
ALT_FILL    = PatternFill("solid", start_color=ALT_BLUE)
WHITE_FILL  = PatternFill("solid", start_color=WHITE)

H_FONT   = Font(name="Arial", bold=True, color=WHITE,  size=11)
D_FONT   = Font(name="Arial",            color="000000", size=10)
T_FONT   = Font(name="Arial", bold=True, color=NAVY,   size=11)
BIG_FONT = Font(name="Arial", bold=True, color=WHITE,  size=15)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# DATE PARSING
# ---------------------------------------------------------------------------

_ORDINAL_RE = re.compile(r"(\d+)(st|nd|rd|th)\b", re.IGNORECASE)

def _strip_ordinals(text):
    """'2nd Mar 2026' -> '2 Mar 2026'"""
    return _ORDINAL_RE.sub(r"\1", str(text))

def _parse_date_series(series):
    """
    Robustly parse a pandas Series of date-like strings.
    Handles ordinals, multiple delimiters, mixed formats.
    """
    cleaned = series.astype(str).apply(_strip_ordinals)
    # Try pandas inference first (covers ISO, DMY, MDY, named months …)
    parsed = pd.to_datetime(cleaned, infer_datetime_format=True,
                            dayfirst=True, errors="coerce")
    # For any that failed, try a set of explicit formats
    _FORMATS = [
        "%d %b %Y", "%d %B %Y", "%d-%b-%Y", "%d-%B-%Y",
        "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d.%m.%Y",
        "%d %b %y", "%d %B %y",
    ]
    mask = parsed.isna()
    if mask.any():
        for fmt in _FORMATS:
            still_bad = parsed.isna()
            if not still_bad.any():
                break
            parsed[still_bad] = pd.to_datetime(
                cleaned[still_bad], format=fmt, errors="coerce"
            )
    return parsed

def _prompt_date(label):
    """Ask user for a date and return a pd.Timestamp. Accepts any sensible format."""
    while True:
        raw = input(f"  Enter {label} (e.g. 1 Jan 2024 / 01/01/2024 / 2024-01-01): ").strip()
        cleaned = _strip_ordinals(raw)
        for fmt in [
            "%d %b %Y", "%d %B %Y", "%d-%b-%Y", "%d-%B-%Y",
            "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d.%m.%Y",
            "%d %b %y", "%d %B %y",
        ]:
            try:
                return pd.Timestamp(datetime.strptime(cleaned, fmt))
            except ValueError:
                pass
        try:
            return pd.Timestamp(pd.to_datetime(cleaned, dayfirst=True))
        except Exception:
            print(f"  Could not understand '{raw}'. Please try again.")

# ---------------------------------------------------------------------------
# STEP 1 — LOAD
# ---------------------------------------------------------------------------

def load_all_sheets(path):
    if not os.path.exists(path):
        sys.exit(f"\nERROR: Input file not found -> {path}\n")
    print(f"\nLoading: {path}")
    return pd.read_excel(path, sheet_name=None, dtype=str)

# ---------------------------------------------------------------------------
# STEP 2 — PROCESS
# ---------------------------------------------------------------------------

def process_sheet(df, start_dt, end_dt):
    """
    Returns (full_df_with_dates_parsed, filtered_and_deduped_df).
    Sheets missing the closure-date column return an empty filtered df.
    """
    df = df.copy()

    if COL_CLOSURE_DATE not in df.columns:
        return df, pd.DataFrame(columns=df.columns)

    df[COL_CLOSURE_DATE] = _parse_date_series(df[COL_CLOSURE_DATE])

    mask     = (df[COL_CLOSURE_DATE] >= start_dt) & (df[COL_CLOSURE_DATE] <= end_dt)
    filtered = df[mask].copy()

    if COL_INCIDENT_ID in filtered.columns:
        filtered = filtered.drop_duplicates(subset=[COL_INCIDENT_ID])

    return df, filtered

# ---------------------------------------------------------------------------
# STEP 3 — AGGREGATE
# ---------------------------------------------------------------------------

def aggregate(counts):
    """
    counts = {sheet_name: int}
    Returns a DataFrame that includes ONLY modules with at least 1 incident.
    """
    rows = [
        {"Module Name": k, "Unique Closed Incidents": v}
        for k, v in counts.items()
        if v > 0
    ]
    if not rows:
        sys.exit("\nNo incidents found in the specified date range across any sheet.\n")
    return pd.DataFrame(rows)

# ---------------------------------------------------------------------------
# STEP 4a — SUMMARY SHEET
# ---------------------------------------------------------------------------

def _write_summary_table(ws, summary_df, start_row):
    ws.row_dimensions[start_row].height = 24
    c = ws.cell(row=start_row, column=1,
                value="Module-wise Unique Closed Incidents")
    c.font      = Font(name="Arial", bold=True, size=13, color=NAVY)
    c.alignment = LEFT
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=3)

    header_row = start_row + 1
    ws.row_dimensions[header_row].height = 22
    for col, text in enumerate(["#", "Module Name", "Unique Closed Incidents"], 1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font      = H_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = _border()

    n = len(summary_df)
    for i, row in enumerate(summary_df.itertuples(index=False), 1):
        r    = header_row + i
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.row_dimensions[r].height = 18
        d    = row._asdict()

        for col_idx, val in enumerate(
            [i, d["Module Name"], d["Unique Closed Incidents"]], 1
        ):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font      = D_FONT
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = LEFT if col_idx == 2 else CENTER

    total_row = header_row + n + 1
    ws.row_dimensions[total_row].height = 22

    for col in [1, 2, 3]:
        cell = ws.cell(row=total_row, column=col)
        cell.fill   = TOTAL_FILL
        cell.border = _thick_border()
        cell.font   = T_FONT
        cell.alignment = CENTER

    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row,   end_column=2)
    ws.cell(row=total_row, column=3,
            value=f"=SUM(C{header_row+1}:C{total_row-1})")

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = max(
        summary_df["Module Name"].astype(str).str.len().max() + 6, 24
    )
    ws.column_dimensions["C"].width = 28

    return header_row, total_row


def _bar_chart(ws, summary_df, header_row):
    n   = len(summary_df)
    bar = BarChart()
    bar.type         = "col"
    bar.grouping     = "clustered"
    bar.title        = "Unique Closed Incidents by Module"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Module"
    bar.style        = 10
    # Scale width with number of modules, min 20, max 34
    bar.width  = min(max(n * 3, 20), 34)
    bar.height = 14

    data = Reference(ws, min_col=3, min_row=header_row,
                     max_col=3, max_row=header_row + n)
    cats = Reference(ws, min_col=2, min_row=header_row + 1,
                     max_row=header_row + n)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)

    series = bar.series[0]
    for idx in range(n):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = _color(idx)
        series.dPt.append(pt)

    return bar


def _pie_chart(ws, summary_df, header_row):
    n   = len(summary_df)
    pie = PieChart()
    pie.title  = "Incident Share by Module"
    pie.style  = 10
    pie.width  = 20
    pie.height = 14

    data = Reference(ws, min_col=3, min_row=header_row,
                     max_col=3, max_row=header_row + n)
    cats = Reference(ws, min_col=2, min_row=header_row + 1,
                     max_row=header_row + n)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)

    series = pie.series[0]
    for idx in range(n):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = _color(idx)
        series.dPt.append(pt)

    return pie


def build_summary_sheet(wb, summary_df, start_dt, end_dt):
    ws = wb.create_sheet(SUMMARY_SHEET_NAME, 0)

    date_range_str = (
        f"{start_dt.strftime('%d %b %Y')}  ->  {end_dt.strftime('%d %b %Y')}"
    )

    # Banner
    ws.row_dimensions[1].height = 42
    banner = ws.cell(row=1, column=1,
                     value=f"Incident Summary Dashboard  |  {date_range_str}")
    banner.font      = BIG_FONT
    banner.fill      = HEADER_FILL
    banner.alignment = CENTER
    ws.merge_cells("A1:P1")

    header_row, total_row = _write_summary_table(ws, summary_df, start_row=3)

    chart_row = total_row + 2

    bar = _bar_chart(ws, summary_df, header_row)
    ws.add_chart(bar, f"A{chart_row}")

    # Place pie chart to the right of the bar chart
    # Bar is ~bar.width/7.5 columns wide; offset by that + 1
    pie_col_idx = round(bar.width / 7.5) + 2
    pie_anchor  = f"{get_column_letter(pie_col_idx)}{chart_row}"
    pie = _pie_chart(ws, summary_df, header_row)
    ws.add_chart(pie, pie_anchor)

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# STEP 4b — DATA SHEETS
# ---------------------------------------------------------------------------

def _write_data_sheet(ws, df, start_dt, end_dt):
    """
    Writes ALL rows to the sheet with styled headers.
    AutoFilter is pre-set to the date range so filtered incidents are visible
    by default; clearing the filter reveals every row.
    """
    if df.empty:
        return

    headers = list(df.columns)
    n_cols  = len(headers)

    ws.row_dimensions[1].height = 20
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill      = SUBHDR_FILL
        cell.alignment = CENTER
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(str(h)) + 4, 14
        )

    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        ws.row_dimensions[row_idx].height = 16
        for col_idx, val in enumerate(row_data, 1):
            try:
                if pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = D_FONT
            cell.fill      = fill
            cell.alignment = LEFT
            cell.border    = _border()

    ws.freeze_panes = "A2"

    # AutoFilter on all columns
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_col}{len(df)+1}"

    # Pre-set date range filter on the closure date column (if present)
    if COL_CLOSURE_DATE in headers:
        col_id_0 = headers.index(COL_CLOSURE_DATE)   # 0-based colId
        cf1 = CustomFilter(
            operator="greaterThanOrEqual",
            val=start_dt.strftime("%Y-%m-%dT%H:%M:%S")
        )
        cf2 = CustomFilter(
            operator="lessThanOrEqual",
            val=end_dt.strftime("%Y-%m-%dT%H:%M:%S")
        )
        fc = FilterColumn(colId=col_id_0)
        fc.customFilters = CustomFilters(customFilter=[cf1, cf2], _and=True)
        ws.auto_filter.filterColumn.append(fc)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    # Prompt for date range
    print("\n" + "="*60)
    print("  INCIDENT REPORT GENERATOR")
    print("="*60)
    print("\nPlease enter the date range for filtering closed incidents.")
    start_dt = _prompt_date("START date (inclusive)")
    end_dt   = _prompt_date("END   date (inclusive)")

    if end_dt < start_dt:
        start_dt, end_dt = end_dt, start_dt
        print("  (Dates swapped — start was after end.)")

    print(f"\n  Range: {start_dt.strftime('%d %b %Y')} → {end_dt.strftime('%d %b %Y')}\n")

    # Step 1 — load
    raw_sheets  = load_all_sheets(INPUT_FILE_PATH)
    sheet_names = list(raw_sheets.keys())
    print(f"Sheets found ({len(sheet_names)}): {sheet_names}\n")

    # Step 2 — process
    processed_raw = {}
    counts        = {}

    for name in sheet_names:
        raw_df, filtered_df = process_sheet(raw_sheets[name], start_dt, end_dt)
        processed_raw[name] = raw_df
        counts[name]        = len(filtered_df)
        status = f"{counts[name]} unique incident(s) in range" if counts[name] else "no incidents in range"
        print(f"  [{name}]  total rows = {len(raw_df)}  |  {status}")

    # Step 3 — aggregate (only modules with ≥1 incident go into the charts)
    summary_df  = aggregate(counts)
    grand_total = summary_df["Unique Closed Incidents"].sum()
    active      = len(summary_df)
    skipped     = len(sheet_names) - active

    print(f"\n  Modules with incidents : {active}")
    if skipped:
        print(f"  Modules with zero hits  : {skipped} (included in data sheets, excluded from charts)")
    print(f"  Grand total             : {grand_total}")

    # Step 4 — build fresh output workbook
    wb = Workbook()
    wb.remove(wb.active)   # remove openpyxl's default blank sheet

    # Summary dashboard (first tab)
    build_summary_sheet(wb, summary_df, start_dt, end_dt)

    # All original data sheets in original order
    for name in sheet_names:
        ws = wb.create_sheet(title=name)
        _write_data_sheet(ws, processed_raw[name], start_dt, end_dt)

    wb.save(OUTPUT_FILE_PATH)
    print(f"\nReport saved -> {OUTPUT_FILE_PATH}")
    print("Done.\n")


if __name__ == "__main__":
    main()
